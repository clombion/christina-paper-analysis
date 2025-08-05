import marimo

__generated_with = "0.14.16"
app = marimo.App(width="medium", app_title="")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# 1. Exploration""")
    return


@app.cell
def _():
    import polars as pl
    from statsmodels.stats.proportion import proportion_confint
    import openpyxl
    return openpyxl, pl, proportion_confint


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell imports the three external Python libraries required for the analysis.

    *   `polars` is imported with the conventional alias `pl`.
    
    *   `proportion_confint` is a specific function imported from the `statsmodels` library.
    
    *   `openpyxl` is imported to handle Excel files.
    

    **Why we do it:**

    *   **`polars`**: This is our primary data manipulation library. All data from the Excel files is read into `polars` DataFrames, which we then use for filtering, grouping, and cleaning.
    
    *   **`proportion_confint`**: This function is central to the analysis. It performs the statistical calculation for the 95% confidence intervals for the PPA and NPA metrics, directly addressing the reviewer's main statistical request.
    
    *   **`openpyxl`**: This library is a necessary dependency for `polars` to read and parse `.xlsx` files. We also use it directly to programmatically list all the sheet names in the workbook, which was a crucial step in understanding the file's structure.
    """
    )
    return


@app.cell(hide_code=True)
def _(openpyxl):
    excel_file_path = 'data/Datasett_MeMed_AUG_2025 – Kopi.xlsx'

    # Load the workbook to inspect it
    workbook = openpyxl.load_workbook(excel_file_path)

    # Get the list of all sheet names
    all_sheet_names = workbook.sheetnames

    print("Found the following sheets in the Excel file:")
    print(all_sheet_names)
    return all_sheet_names, excel_file_path


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell uses the `openpyxl` library to open the Excel workbook from the specified file path. It then extracts a list of all the worksheet names contained within that file and stores them in the `all_sheet_names` variable. Finally, it prints this list to the console.

    **Why we do it:** We do this to understand the structure of the multi-sheet Excel file without manual inspection. It allowed us to identify our target raw data sheet (`'full dataset'`). We also realized that the file also contained many pre-filtered and metadata sheets.
    """
    )
    return


@app.cell(hide_code=True)
def _(all_sheet_names, excel_file_path, pl):
    # The list 'all_sheet_names' from the previous cell is used here
    for sheet in all_sheet_names:
        print("="*50)
        print(f"Previewing sheet: '{sheet}'")
        print("="*50)

        # Load the current sheet
        # Note: This might raise an error on metadata sheets if they aren't table-like.
        temp_df = pl.read_excel(
            source=excel_file_path,
            sheet_name=sheet,
            infer_schema_length=None
        )
    
        # Print the first 10 rows
        print(temp_df.head(10))
        print("\n")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell iterates through the `all_sheet_names` list. For each sheet name in the list, it loads the corresponding worksheet into a temporary `polars` DataFrame and prints the name of the sheet followed by its first 10 rows. The `infer_schema_length=None` argument is used to help `polars` correctly guess the data type of each column by scanning the entire sheet, which reduces data type warnings.

    **Why we do it:** We do this to efficiently survey the contents of all 23 sheets in the Excel file. This automated preview is faster than manual inspection and provides a complete overview of the workbook's structure. The output helps distinguish between raw data sheets, metadata, and pre-calculated summary tables.
    """
    )
    return


@app.cell(hide_code=True)
def _(pl):
    sheet_name_to_load = 'full dataset'

    df = pl.read_excel(
        source='data/Datasett_MeMed_AUG_2025 – Kopi.xlsx',
        sheet_name=sheet_name_to_load,
        infer_schema_length=None
    )

    df
    return (df,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell loads a single worksheet, named `'full dataset'`, from the Excel workbook into a `polars` DataFrame called `df`, and previews it.

    **Why we do it:** We do this to load the primary data source for the analysis into memory. The previous exploration identified the `'full dataset'` sheet as the one containing the complete raw data. This step prepares the main `df` variable that is used in all subsequent filtering and calculation steps.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# 2. Analysis""")
    return


@app.cell(hide_code=True)
def _(df):
    # See the unique values and their counts for 'Oppst AB 48t'
    print("Value counts for 'Oppst AB 48t':")
    print(df["Oppst AB 48t"].value_counts())

    print("\n" + "="*40 + "\n")

    # See the unique values and their counts for 'AB 72t'
    print("Value counts for 'AB 72t':")
    print(df["AB 72t"].value_counts())
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell inspects two columns from the main DataFrame `df`: `'Oppst AB 48t'` and `'AB 72t'`. For each column, it calculates and prints a summary table showing every unique value and the number of times it appears.

    **Why we do it:** We do this to understand how the data is encoded in these key columns. A previous error indicated a data type mismatch, so this step was necessary to see the actual values used (e.g., numbers like `0, 1, 2` vs. text like `'JA', 'NEI'`).
    """
    )
    return


@app.cell(hide_code=True)
def _(df, pl):
    # --- Define the clinical management cohort based on the paper's rules ---

    # 1. Patients managed as "viral/non-bacterial"
    viral_management_group = df.filter(
        (pl.col("healthy_control") == 0) &
        (pl.col("AB 72t") == "NEI")
    )

    # 2. Patients managed as "bacterial"
    bacterial_management_group = df.filter(
        (pl.col("healthy_control") == 0) &
        (pl.col("AB 72t") == "JA") &
        (pl.col("Oppst AB 48t") == 1) # Must have started AB within 48h
    )

    # 3. Combine them to get the final cohort of 442 patients
    clinical_df = pl.concat([
        viral_management_group,
        bacterial_management_group
    ])

    # --- Create the contingency table ---
    clinical_contingency = clinical_df.group_by("AB 72t", "MeMed score category").agg(
        pl.len().alias("count")
    ).sort("AB 72t", descending=True)

    # For clarity, rename the values to match our previous analysis
    clinical_contingency = clinical_contingency.with_columns(
        pl.when(pl.col("AB 72t") == "JA")
          .then(pl.lit("Bacterial management"))
          .otherwise(pl.lit("Viral/non-bacterial management"))
          .alias("Clinical_management")
    ).drop("AB 72t")


    print("--- Clinical Management Cohort ---")
    print(f"Filtered down to {clinical_df.height} patients.")
    print("Contingency Table:")
    print(clinical_contingency)
    return (clinical_contingency,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell constructs the 442-patient clinical management cohort and the final contingency table in several steps:

    *   **1\. Filters for the "Viral/Non-bacterial" Group (n=93):**
    
        *   It selects rows from the main `df` where:
        
            *   The `healthy_control` column is `0` (to exclude the control group).
            
            *   The `AB 72t` column is `"NEI"` (indicating antibiotic treatment was less than 72 hours).
            
    *   **2\. Filters for the "Bacterial" Group (n=349):**
    
        *   It selects rows from the main `df` where:
        
            *   The `healthy_control` column is `0`.
            
            *   The `AB 72t` column is `"JA"` (indicating treatment was 72 hours or more).
            
            *   The `Oppst AB 48t` column is `1` (ensuring antibiotics were also started within the first 48 hours).
            
    *   **3\. Creates the Final Cohort and Contingency Table:**
    
        *   It combines these two separate groups into the final `clinical_df` DataFrame.
        
        *   It then groups this 442-patient cohort by the `AB 72t` and `MeMed score category` columns, counting the number of patients in each subgroup to create the `clinical_contingency` table.
        
        *   Finally, it adds a more descriptive `Clinical_management` column to the table for clarity.
        

    **Why we do it:** We do this to precisely replicate the clinical management cohort from the paper. The multi-step filtering is necessary to apply the study's specific inclusion/exclusion criteria. This logic correctly defines the "bacterial" group as not just receiving long-term antibiotics, but also starting them promptly. This process ensures the resulting contingency table is the correct basis for calculating the PPA and NPA.
    """
    )
    return


@app.cell(hide_code=True)
def _(clinical_contingency, pl, proportion_confint):
    # --- PPA Calculation (Bacterial Management) ---
    # Get the total number of patients in the 'Bacterial management' group
    ppa_total = clinical_contingency.filter(
        pl.col("Clinical_management") == "Bacterial management"
    )["count"].sum()

    # Get the number of those correctly identified with a "Bacterial" score (code = 1)
    ppa_success = clinical_contingency.filter(
        (pl.col("Clinical_management") == "Bacterial management") &
        (pl.col("MeMed score category") == 1)
    )["count"][0]

    # Calculate the confidence interval
    ppa_ci = proportion_confint(count=ppa_success, nobs=ppa_total, alpha=0.05, method='wilson')

    print(f"PPA (Bacterial Management): {ppa_success / ppa_total:.2%} (95% CI: {ppa_ci[0]:.2%} - {ppa_ci[1]:.2%})")


    # --- NPA Calculation (Viral/Non-bacterial Management) ---
    # Get the total number of patients in the 'Viral/non-bacterial management' group
    npa_total = clinical_contingency.filter(
        pl.col("Clinical_management") == "Viral/non-bacterial management"
    )["count"].sum()

    # Get the number of those correctly identified with a "Viral" score (code = 2)
    npa_success = clinical_contingency.filter(
        (pl.col("Clinical_management") == "Viral/non-bacterial management") &
        (pl.col("MeMed score category") == 2)
    )["count"][0]

    # Calculate the confidence interval
    npa_ci = proportion_confint(count=npa_success, nobs=npa_total, alpha=0.05, method='wilson')

    print(f"NPA (Viral/Non-bacterial Management): {npa_success / npa_total:.2%} (95% CI: {npa_ci[0]:.2%} - {npa_ci[1]:.2%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell calculates the Positive and Negative Percent Agreement (PPA/NPA) and their 95% confidence intervals from the `clinical_contingency` table.

    *   **For the PPA (Bacterial Management):**
    
        *   It determines the total number of patients where `Clinical_management` is `"Bacterial management"` (n=349).
        
        *   It finds the number of those patients who also had a 'Bacterial' MeMed score (`MeMed score category` is `1`), which represents the successful agreements (n=314).
        
        *   It uses `proportion_confint` to calculate the 95% confidence interval for this proportion.
        
    *   **For the NPA (Viral/Non-bacterial Management):**
    
        *   It determines the total number of patients where `Clinical_management` is `"Viral/non-bacterial management"` (n=93).
        
        *   It finds the number of those patients who also had a 'Viral' MeMed score (`MeMed score category` is `2`), representing the successful agreements (n=41).
        
        *   It uses `proportion_confint` to calculate the 95% confidence interval for this proportion.
        

    **Why we do it:** This cell performs the core statistical analysis for the clinical cohort to meet the reviewer's requirements. The PPA measures the test's agreement with clinical decisions for bacterial cases, while the NPA measures its agreement for non-bacterial cases. The primary goal is to compute the confidence intervals, which quantify the statistical uncertainty of these performance metrics.
    """
    )
    return


@app.cell(hide_code=True)
def _(df, pl):
    # Filter for the 370 patients in the molecular analysis cohort
    # The RTi Category column codes are: 1=Bacterial, 2=Viral, 3=No Detection
    molecular_df = df.filter(
        (pl.col("healthy_control") == 0) &
        (pl.col("RTi Category  FAP B or V (BV=B)").is_in([1, 2, 3]))
    )

    # Create the contingency table for the PPA/NPA analysis (groups 1 and 2)
    molecular_contingency = molecular_df.filter(
        pl.col("RTi Category  FAP B or V (BV=B)").is_in([1, 2])
    ).group_by("RTi Category  FAP B or V (BV=B)", "MeMed score category").agg(
        pl.len().alias("count")
    ).sort("RTi Category  FAP B or V (BV=B)")

    # Map the numeric codes to text for clarity
    molecular_contingency = molecular_contingency.with_columns(
        pl.when(pl.col("RTi Category  FAP B or V (BV=B)") == 1)
          .then(pl.lit("Bacterial detections"))
          .otherwise(pl.lit("Viral detections"))
          .alias("Molecular_Detection")
    ).drop("RTi Category  FAP B or V (BV=B)")

    print("--- Molecular Detection Cohort ---")
    print(f"Filtered down to {molecular_df.height} patients.")
    print("Final Contingency Table:")
    print(molecular_contingency)
    return (molecular_contingency,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **What this cell does:** This cell prepares the data and creates the contingency table for the molecular cohort in several steps:

    *   **1\. Defines the Molecular Cohort (`molecular_df`):**
    
        *   It filters the main `df` DataFrame to select all non-control patients (where `healthy_control` is `0`).
        
        *   It includes only patients who have a valid molecular test result by keeping rows where `RTi Category FAP B or V (BV=B)` is `1` (Bacterial), `2` (Viral), or `3` (No Detection).
        
    *   **2\. Creates the Contingency Table (`molecular_contingency`):**
    
        *   It takes the `molecular_df` from the step above and applies a second filter to keep only the 'Bacterial' (code `1`) and 'Viral' (code `2`) groups, because the 'No Detection' group is not part of the PPA/NPA analysis.
        
        *   It then groups this data by the molecular result and the `MeMed score category`, and counts the number of patients in each subgroup.
        
        *   Finally, it adds a descriptive `Molecular_Detection` column to the table for clarity.
        

    **Why we do it:** We do this to prepare the data for calculating the PPA and NPA for the molecular detection cohort. The filtering steps are necessary to isolate the specific patient population relevant to this analysis, as defined by the study's methodology. The resulting contingency table aggregates the data into the required format for the statistical calculations in the final step.
    """
    )
    return


@app.cell(hide_code=True)
def _(molecular_contingency, pl, proportion_confint):
    # --- PPA Calculation (Bacterial Detections) ---
    ppa_total_mol = molecular_contingency.filter(
        pl.col("Molecular_Detection") == "Bacterial detections"
    )["count"].sum()

    ppa_success_mol = molecular_contingency.filter(
        (pl.col("Molecular_Detection") == "Bacterial detections") &
        (pl.col("MeMed score category") == 1) # Bacterial score
    )["count"][0]

    ppa_ci_mol = proportion_confint(count=ppa_success_mol, nobs=ppa_total_mol, alpha=0.05, method='wilson')
    print(f"PPA (Bacterial Detections): {ppa_success_mol / ppa_total_mol:.2%} (95% CI: {ppa_ci_mol[0]:.2%} - {ppa_ci_mol[1]:.2%})")


    # --- NPA Calculation (Viral Detections) ---
    npa_total_mol = molecular_contingency.filter(
        pl.col("Molecular_Detection") == "Viral detections"
    )["count"].sum()

    npa_success_mol = molecular_contingency.filter(
        (pl.col("Molecular_Detection") == "Viral detections") &
        (pl.col("MeMed score category") == 2) # Viral score
    )["count"][0]

    npa_ci_mol = proportion_confint(count=npa_success_mol, nobs=npa_total_mol, alpha=0.05, method='wilson')
    print(f"NPA (Viral Detections): {npa_success_mol / npa_total_mol:.2%} (95% CI: {npa_ci_mol[0]:.2%} - {npa_ci_mol[1]:.2%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell calculates the PPA and NPA and their 95% confidence intervals using the `molecular_contingency` table.

    *   **For the PPA (Bacterial Detections):**
    
        *   It determines the total number of patients in the `"Bacterial detections"` group.
        
        *   It finds the number of those patients who also had a 'Bacterial' MeMed score (`MeMed score category` is `1`).
        
        *   It uses `proportion_confint` to calculate the 95% confidence interval for this proportion.
        
    *   **For the NPA (Viral Detections):**
    
        *   It determines the total number of patients in the `"Viral detections"` group.
        
        *   It finds the number of those patients who also had a 'Viral' MeMed score (`MeMed score category` is `2`).
        
        *   It uses `proportion_confint` to calculate the 95% confidence interval for this proportion.
        

    **Why we do it:** This cell performs the core statistical analysis for the molecular cohort, which is the final set of results needed to address the reviewer's feedback. Here, the PPA and NPA measure the test's agreement against molecular pathogen detection. This provides a different perspective on the test's performance compared to the clinical management benchmark and completes the primary goal of the analysis.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- Manually create the contingency table from the paper's Table 5 ---
    # Bacterial Detections (n=246): 217 Bacterial, 18 Viral, 11 Equivocal
    # Viral Detections (n=73): 40 Bacterial, 22 Viral, 11 Equivocal

    # --- PPA Calculation (Bacterial Detections) ---
    ppa_success_paper = 217
    ppa_total_paper = 246
    ppa_ci_paper = proportion_confint(count=ppa_success_paper, nobs=ppa_total_paper, alpha=0.05, method='wilson')
    print("--- CIs for the Exact Numbers in Paper's Table 5 ---")
    print(f"PPA (Bacterial Detections): {ppa_success_paper / ppa_total_paper:.2%} (95% CI: {ppa_ci_paper[0]:.2%} - {ppa_ci_paper[1]:.2%})")


    # --- NPA Calculation (Viral Detections) ---
    npa_success_paper = 22
    npa_total_paper = 73
    npa_ci_paper = proportion_confint(count=npa_success_paper, nobs=npa_total_paper, alpha=0.05, method='wilson')
    print(f"NPA (Viral Detections): {npa_success_paper / npa_total_paper:.2%} (95% CI: {npa_ci_paper[0]:.2%} - {npa_ci_paper[1]:.2%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell directly calculates the PPA and NPA confidence intervals for the molecular cohort using hard-coded numbers taken from the paper's Table 5.

    *   **For the PPA (Bacterial Detections):**
    
        *   It sets the number of successful agreements to `217` and the total number of patients to `246`, matching the paper's data for this group.
        
        *   It then computes and prints the PPA and its 95% confidence interval.
        
    *   **For the NPA (Viral Detections):**
    
        *   It sets the number of successful agreements to `22` and the total to `73`, matching the paper's data.
        
        *   It then computes and prints the NPA and its 95% confidence interval.
        

    **Why we do it:** We do this to provide the most direct and accurate answer to the reviewer. Our previous steps showed that programmatically filtering the raw dataset to perfectly replicate the final 370-patient molecular cohort was challenging. This final step sidesteps those data cleaning issues by using the adjudicated numbers directly from the paper's table, ensuring the confidence intervals correspond exactly to the data presented in the manuscript.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# 3. Comparison with and without equivocals""")
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- PPA Calculation (Equivocals Included) ---
    # TP = 314, FN = 35 (18 viral + 17 equivocal)
    ppa_total_with_eq = 314 + 35
    ppa_success_with_eq = 314
    ppa_ci_with_eq = proportion_confint(count=ppa_success_with_eq, nobs=ppa_total_with_eq, alpha=0.05, method='wilson')

    # --- NPA Calculation (Equivocals Included) ---
    # TN = 41, FP = 52 (35 bacterial + 17 equivocal)
    npa_total_with_eq = 41 + 52
    npa_success_with_eq = 41
    npa_ci_with_eq = proportion_confint(count=npa_success_with_eq, nobs=npa_total_with_eq, alpha=0.05, method='wilson')


    print("--- CIs for Clinical Cohort (Including Equivocals, n=442) ---")
    print(f"PPA: {ppa_success_with_eq / ppa_total_with_eq:.2%} (95% CI: {ppa_ci_with_eq[0]:.2%} - {ppa_ci_with_eq[1]:.2%})")
    print(f"NPA: {npa_success_with_eq / npa_total_with_eq:.2%} (95% CI: {npa_ci_with_eq[0]:.2%} - {npa_ci_with_eq[1]:.2%})")
    return (
        npa_ci_with_eq,
        npa_success_with_eq,
        npa_total_with_eq,
        ppa_ci_with_eq,
        ppa_success_with_eq,
        ppa_total_with_eq,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    #### Explanation

    **What this cell does:** This cell calculates the PPA, NPA, and their 95% confidence intervals for the full clinical management cohort (n=442).

    *   **For the PPA:** It sets the True Positives (TP) to `314` and False Negatives (FN) to `35`. This FN count combines patients with both viral scores (18) and equivocal scores (17).
    
    *   **For the NPA:** It sets the True Negatives (TN) to `41` and False Positives (FP) to `52`. This FP count combines patients with both bacterial scores (35) and equivocal scores (17).
    

    **Why we do it:** This is the same primary analysis as done previously for the clinical cohort. In this scenario, an "equivocal" test result is treated as a test failure (i.e., it does not agree with the clinical management decision). This provides a measure of the test's performance in a real-world setting where non-definitive results must be handled.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- PPA Calculation (Equivocals Excluded) ---
    # TP = 314, FN = 18
    ppa_total_no_eq = 314 + 18
    ppa_success_no_eq = 314
    ppa_ci_no_eq = proportion_confint(count=ppa_success_no_eq, nobs=ppa_total_no_eq, alpha=0.05, method='wilson')

    # --- NPA Calculation (Equivocals Excluded) ---
    # TN = 41, FP = 35
    npa_total_no_eq = 41 + 35
    npa_success_no_eq = 41
    npa_ci_no_eq = proportion_confint(count=npa_success_no_eq, nobs=npa_total_no_eq, alpha=0.05, method='wilson')


    print("\n--- CIs for Clinical Cohort (Excluding Equivocals, n=408) ---")
    print(f"PPA: {ppa_success_no_eq / ppa_total_no_eq:.2%} (95% CI: {ppa_ci_no_eq[0]:.2%} - {ppa_ci_no_eq[1]:.2%})")
    print(f"NPA: {npa_success_no_eq / npa_total_no_eq:.2%} (95% CI: {npa_ci_no_eq[0]:.2%} - {npa_ci_no_eq[1]:.2%})")
    return (
        npa_ci_no_eq,
        npa_success_no_eq,
        npa_total_no_eq,
        ppa_ci_no_eq,
        ppa_success_no_eq,
        ppa_total_no_eq,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    #### Explanation

    **What this cell does:** This cell recalculates the PPA, NPA, and their confidence intervals for a subset of the clinical cohort (n=408) that excludes the 34 patients who had equivocal test results.

    *   **For the PPA:** The FN count is now `18` (only patients with a viral score).
    
    *   **For the NPA:** The FP count is now `35` (only patients with a bacterial score).
    

    **Why we do it:** This is a sensitivity analysis. Its purpose is to evaluate the test's performance based only on cases where it produced a definitive (bacterial or viral) result. Comparing these results to the primary analysis allows us to quantify the impact of the equivocal results.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- PPA Calculation (Equivocals Included) ---
    # From Table 5: TP = 217, FN = 18 (viral score) + 11 (equivocal score) = 29
    ppa_total_with_eq_mol = 217 + 29
    ppa_success_with_eq_mol = 217
    ppa_ci_with_eq_mol = proportion_confint(count=ppa_success_with_eq_mol, nobs=ppa_total_with_eq_mol, alpha=0.05, method='wilson')

    # --- NPA Calculation (Equivocals Included) ---
    # From Table 5: TN = 22, FP = 40 (bacterial score) + 11 (equivocal score) = 51
    npa_total_with_eq_mol = 22 + 51
    npa_success_with_eq_mol = 22
    npa_ci_with_eq_mol = proportion_confint(count=npa_success_with_eq_mol, nobs=npa_total_with_eq_mol, alpha=0.05, method='wilson')


    print("--- CIs for Molecular Cohort (Including Equivocals) ---")
    print(f"PPA (Bacterial Detections): {ppa_success_with_eq_mol / ppa_total_with_eq_mol:.2%} (95% CI: {ppa_ci_with_eq_mol[0]:.2%} - {ppa_ci_with_eq_mol[1]:.2%})")
    print(f"NPA (Viral Detections): {npa_success_with_eq_mol / npa_total_with_eq_mol:.2%} (95% CI: {npa_ci_with_eq_mol[0]:.2%} - {npa_ci_with_eq_mol[1]:.2%})")
    return (
        npa_ci_with_eq_mol,
        npa_success_with_eq_mol,
        npa_total_with_eq_mol,
        ppa_ci_with_eq_mol,
        ppa_success_with_eq_mol,
        ppa_total_with_eq_mol,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    #### Explanation

    **What this cell does:** This cell calculates the PPA, NPA, and their 95% confidence intervals for the molecular detection cohort, using the numbers from the paper's Table 5.

    *   **For the PPA:** The FN count of `29` combines patients with viral scores (18) and equivocal scores (11).
    
    *   **For the NPA:** The FP count of `51` combines patients with bacterial scores (40) and equivocal scores (11).
    

    **Why we do it:** This is the same primary analysis as done previously for the molecular cohort. An "equivocal" result is treated as a test failure against the molecular benchmark. This provides a measure of the test's performance against pathogen detection for all patients in this cohort.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- PPA Calculation (Equivocals Excluded) ---
    # From Table 5: TP = 217, FN = 18 (viral score only)
    ppa_total_no_eq_mol = 217 + 18
    ppa_success_no_eq_mol = 217
    ppa_ci_no_eq_mol = proportion_confint(count=ppa_success_no_eq_mol, nobs=ppa_total_no_eq_mol, alpha=0.05, method='wilson')

    # --- NPA Calculation (Equivocals Excluded) ---
    # From Table 5: TN = 22, FP = 40 (bacterial score only)
    npa_total_no_eq_mol = 22 + 40
    npa_success_no_eq_mol = 22
    npa_ci_no_eq_mol = proportion_confint(count=npa_success_no_eq_mol, nobs=npa_total_no_eq_mol, alpha=0.05, method='wilson')


    print("\n--- CIs for Molecular Cohort (Excluding Equivocals) ---")
    print(f"PPA (Bacterial Detections): {ppa_success_no_eq_mol / ppa_total_no_eq_mol:.2%} (95% CI: {ppa_ci_no_eq_mol[0]:.2%} - {ppa_ci_no_eq_mol[1]:.2%})")
    print(f"NPA (Viral Detections): {npa_success_no_eq_mol / npa_total_no_eq_mol:.2%} (95% CI: {npa_ci_no_eq_mol[0]:.2%} - {npa_ci_no_eq_mol[1]:.2%})")
    return (
        npa_ci_no_eq_mol,
        npa_success_no_eq_mol,
        npa_total_no_eq_mol,
        ppa_ci_no_eq_mol,
        ppa_success_no_eq_mol,
        ppa_total_no_eq_mol,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    #### Explanation

    **What this cell does:** This cell recalculates the PPA, NPA, and their confidence intervals for the subset of the molecular cohort that excludes patients with equivocal test results.

    *   **For the PPA:** The FN count is now `18`.
    
    *   **For the NPA:** The FP count is now `40`.
    

    **Why we do it:** This is the corresponding sensitivity analysis for the molecular cohort. It assesses the test's performance only on cases with a definitive result against the molecular benchmark, allowing for a direct comparison to the primary analysis to understand the impact of equivocal results in this context.
    """
    )
    return


@app.cell(hide_code=True)
def _(
    npa_ci_no_eq,
    npa_ci_no_eq_mol,
    npa_ci_with_eq,
    npa_ci_with_eq_mol,
    npa_success_no_eq,
    npa_success_no_eq_mol,
    npa_success_with_eq,
    npa_success_with_eq_mol,
    npa_total_no_eq,
    npa_total_no_eq_mol,
    npa_total_with_eq,
    npa_total_with_eq_mol,
    pl,
    ppa_ci_no_eq,
    ppa_ci_no_eq_mol,
    ppa_ci_with_eq,
    ppa_ci_with_eq_mol,
    ppa_success_no_eq,
    ppa_success_no_eq_mol,
    ppa_success_with_eq,
    ppa_success_with_eq_mol,
    ppa_total_no_eq,
    ppa_total_no_eq_mol,
    ppa_total_with_eq,
    ppa_total_with_eq_mol,
):
    # This cell assumes the previous calculation cells have been run and their
    # variables (ppa_ci_with_eq, npa_ci_no_eq_mol, etc.) are in memory.

    # Create a list of dictionaries, programmatically building the strings from our variables
    summary_data = [
        {
            "Cohort": "Clinical Management",
            "Metric": "PPA",
            "Including Equivocals": f"{(ppa_success_with_eq / ppa_total_with_eq):.2%} (95% CI: {ppa_ci_with_eq[0]:.2%} - {ppa_ci_with_eq[1]:.2%})",
            "Excluding Equivocals": f"{(ppa_success_no_eq / ppa_total_no_eq):.2%} (95% CI: {ppa_ci_no_eq[0]:.2%} - {ppa_ci_no_eq[1]:.2%})",
        },
        {
            "Cohort": "Clinical Management",
            "Metric": "NPA",
            "Including Equivocals": f"{(npa_success_with_eq / npa_total_with_eq):.2%} (95% CI: {npa_ci_with_eq[0]:.2%} - {npa_ci_with_eq[1]:.2%})",
            "Excluding Equivocals": f"{(npa_success_no_eq / npa_total_no_eq):.2%} (95% CI: {npa_ci_no_eq[0]:.2%} - {npa_ci_no_eq[1]:.2%})",
        },
        {
            "Cohort": "Molecular Detection",
            "Metric": "PPA",
            "Including Equivocals": f"{(ppa_success_with_eq_mol / ppa_total_with_eq_mol):.2%} (95% CI: {ppa_ci_with_eq_mol[0]:.2%} - {ppa_ci_with_eq_mol[1]:.2%})",
            "Excluding Equivocals": f"{(ppa_success_no_eq_mol / ppa_total_no_eq_mol):.2%} (95% CI: {ppa_ci_no_eq_mol[0]:.2%} - {ppa_ci_no_eq_mol[1]:.2%})",
        },
        {
            "Cohort": "Molecular Detection",
            "Metric": "NPA",
            "Including Equivocals": f"{(npa_success_with_eq_mol / npa_total_with_eq_mol):.2%} (95% CI: {npa_ci_with_eq_mol[0]:.2%} - {npa_ci_with_eq_mol[1]:.2%})",
            "Excluding Equivocals": f"{(npa_success_no_eq_mol / npa_total_no_eq_mol):.2%} (95% CI: {npa_ci_no_eq_mol[0]:.2%} - {npa_ci_no_eq_mol[1]:.2%})",
        },
    ]

    # Create and display the summary DataFrame
    summary_df = pl.DataFrame(summary_data)

    print("--- Comparison of Performance Metrics ---")
    print(summary_df)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **What this cell does:** This cell gathers the results stored in the various `ppa_*` and `npa_*` variables from all the previous calculation steps. It uses f-strings to format these results into readable strings (percentage and confidence interval). Finally, it organizes this formatted data into a `polars` DataFrame to create a clean, side-by-side comparison table.

    **Why we do it:** We do this to create a **dynamic and reproducible** summary of the entire analysis.
    """
    )
    return


@app.cell
def _(mo):
    mo.md(r"""# 4. Confidence Intervals for Descriptive Statistics""")
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # --- For the "Bacterial management" group ---
    n_success_bact = 4
    n_total_bact = 349
    ci_bact = proportion_confint(count=n_success_bact, nobs=n_total_bact, method='wilson')
    print(f"Bacterial Management Group (High Likelihood Viral Score):")
    print(f"Point Estimate: {n_success_bact / n_total_bact:.1%}")
    print(f"95% CI: {ci_bact[0]:.1%} - {ci_bact[1]:.1%}")


    # --- For the "Viral/non-bacterial management" group ---
    n_success_viral = 21
    n_total_viral = 93
    ci_viral = proportion_confint(count=n_success_viral, nobs=n_total_viral, method='wilson')
    print(f"\nViral/Non-bacterial Management Group (High Likelihood Viral Score):")
    print(f"Point Estimate: {n_success_viral / n_total_viral:.1%}")
    print(f"95% CI: {ci_viral[0]:.1%} - {ci_viral[1]:.1%}")

    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Explanation

    **What this cell does:** This cell calculates the 95% confidence intervals for a specific data point presented in the descriptive tables: the percentage of patients with a "High likelihood viral infection" score, stratified by management group.

    *   **For the "Bacterial management" group:**
    
        *   It uses the raw counts from Table 1, where `n=4` out of a total `N=349`.
        
        *   It calls `proportion_confint` to compute the CI for the proportion 4/349.
        
    *   **For the "Viral/non-bacterial management" group:**
    
        *   It uses the raw counts from Table 1, where `n=21` out of a total `N=93`.
        
        *   It computes the CI for the proportion 21/93.
        

    **Why we do it:** We do this to apply the reviewer's feedback on "uncertainty quantification" to the descriptive tables in a thorough manner. While not the primary performance metrics, the rows describing the MeMed test's results in Table 1 and Table 4 can be considered performance-related statistics. This calculation provides the required confidence intervals, making the tables more statistically robust. This code serves as a template that can be repeated for the other relevant rows to fully address the reviewer's comment.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # For Table 3, "Bacterial management" group, "Bacterial [%]" row
    ci_table3 = proportion_confint(count=15, nobs=22, method='wilson')
    print(f"68.2% (95% CI: {ci_table3[0]:.1%} - {ci_table3[1]:.1%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **What this cell does:** This cell calculates the 95% confidence interval for a specific proportion from Table 3. It uses the raw counts of patients who were ultimately diagnosed as non-infectious but were managed as "bacterial" (`N=22`), and of those, the number who received a "bacterial" MeMed score (`n=15`).

    **Why we do it:** We do this to add uncertainty quantification to the subgroup analysis in Table 3. This table assesses the test's performance in patients without a final infectious diagnosis, making it important to quantify the statistical certainty of these results to fully address the reviewer's feedback.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # For Supplementary Table 1, "Moderate likelihood of viral infection" row
    ci_supp1 = proportion_confint(count=13, nobs=20, method='wilson')
    print(f"65.0% (95% CI: {ci_supp1[0]:.1%} - {ci_supp1[1]:.1%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **What this cell does:** This cell calculates the 95% confidence interval for a proportion from Supplementary Table 1. It uses the raw counts of healthy controls (N=20) and the number of those who received a "Moderate likelihood of viral infection" score (n=13).

    **Why we do it:** We do this to add statistical rigor to the data describing the test's behavior in the healthy control group. Providing a CI quantifies the uncertainty around the 65% point estimate, which is a key measure of the test's baseline performance in a non-diseased population.
    """
    )
    return


@app.cell(hide_code=True)
def _(proportion_confint):
    # For Supplementary Table 2, "Bacterial management" group, "Bacterial MeMed BV® test scores [%]" row
    ci_supp2 = proportion_confint(count=55, nobs=60, method='wilson')
    print(f"91.7% (95% CI: {ci_supp2[0]:.1%} - {ci_supp2[1]:.1%})")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **What this cell does:** This cell calculates the 95% confidence interval for a proportion from Supplementary Table 2. It uses the raw counts of patients with "no microbial detections" who were still managed as "bacterial" (`N=60`), and of those, the number who received a "bacterial" MeMed score (`n=55`).

    **Why we do it:** We do this to add uncertainty quantification to the important subgroup of patients where no pathogen could be found. Understanding the confidence in the test's performance in this ambiguous group is critical, and this calculation provides that necessary statistical context.
    """
    )
    return


@app.cell
def _():
    import marimo as mo
    return (mo,)


if __name__ == "__main__":
    app.run()
